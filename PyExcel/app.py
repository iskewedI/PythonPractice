import openpyxl

wb = openpyxl.Workbook()

sheet = wb.create_sheet("Sheet1", 0)

# cell = sheet["a1"]
# cell = sheet["a"]
# cell = sheet["a:c"]
# cell = sheet.cell(row=1, column=1)

for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        cell = sheet.cell(row, column)
        print(cell.value)

# If we call sheet.cell() to get a cell, but the cell DIDN'T exists previously, it will create it!
wb.save("trial.xlsx")
